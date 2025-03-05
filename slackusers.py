import os
import requests
import openpyxl
from openpyxl.styles import Font

def get_slack_users_from_channel(channel_id, slack_token):
    """
    Retrieves all user IDs and names from a Slack channel.

    Args:
        channel_id: The ID of the Slack channel.
        slack_token: Your Slack API token.

    Returns:
        A dictionary where keys are user IDs and values are usernames, or None if an error occurs.
    """
    user_data = {}
    next_cursor = None

    try:
        while True:
            url = f"https://slack.com/api/conversations.members?channel={channel_id}"
            headers = {"Authorization": f"Bearer {slack_token}"}
            if next_cursor:
                url += f"&cursor={next_cursor}"

            response = requests.get(url, headers=headers)
            response.raise_for_status()  # Raise HTTPError for bad responses (4xx or 5xx)
            data = response.json()

            if not data["ok"]:
                print(f"Error getting channel members: {data.get('error')}")
                return None

            user_ids = data["members"]

            for user_id in user_ids:
                user_info_url = f"https://slack.com/api/users.info?user={user_id}"
                user_info_response = requests.get(user_info_url, headers=headers)
                user_info_response.raise_for_status()
                user_info_data = user_info_response.json()

                if user_info_data["ok"]:
                    user_data[user_id] = user_info_data["user"]["name"]
                else:
                    print(f"Error getting user info for {user_id}: {user_info_data.get('error')}")

            next_cursor = data["response_metadata"].get("next_cursor")
            if not next_cursor:
                break

        return user_data

    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
        return None
    except KeyError as e:
        print(f"KeyError: {e}. Check the API response format.")
        return None

def create_excel_sheet(user_data, filename="slack_users.xlsx"):
    """
    Creates an Excel sheet with user IDs and usernames.

    Args:
        user_data: A dictionary where keys are user IDs and values are usernames.
        filename: The name of the Excel file to create.
    """
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Slack Users"

        # Add headers with bold font
        headers = ["User ID", "Username"]
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = Font(bold=True)

        # Add user data
        for row_num, (user_id, username) in enumerate(user_data.items(), 2):
            sheet.cell(row=row_num, column=1).value = user_id
            sheet.cell(row=row_num, column=2).value = username

        workbook.save(filename)
        print(f"Excel sheet '{filename}' created successfully.")

    except Exception as e:
        print(f"An error occurred while creating the Excel sheet: {e}")

# Example usage:
if __name__ == "__main__":
    slack_token = "REPLACE YOUR SLACK TOKEN"
    channel_id = "CEF88GEKA"  # Replace with your channel ID

    if not slack_token:
        print("Error: SLACK_TOKEN environment variable not set.")
    else:
        users = get_slack_users_from_channel(channel_id, slack_token)
        if users:
            create_excel_sheet(users)
